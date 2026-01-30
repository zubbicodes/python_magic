import argparse
import base64
import json
import math
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any


def _json_compatible(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (bytes, bytearray)):
        return {"_type": "bytes", "base64": base64.b64encode(value).decode("ascii")}
    if isinstance(value, (list, tuple)):
        return [_json_compatible(v) for v in value]
    if isinstance(value, dict):
        return {str(k): _json_compatible(v) for k, v in value.items()}
    return str(value)


def _looks_like_header_row(row: list[Any]) -> bool:
    if not row:
        return False
    non_empty = [cell for cell in row if cell is not None]
    if len(non_empty) < 2:
        return False

    normalized: list[str] = []
    for cell in non_empty:
        if not isinstance(cell, str):
            return False
        text = cell.strip()
        if not text:
            return False
        normalized.append(text)
    return len(normalized) == len(set(normalized))


def _make_unique_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for header in headers:
        key = header
        if key not in counts:
            counts[key] = 1
            result.append(key)
            continue

        counts[key] += 1
        result.append(f"{key}_{counts[key]}")
    return result


def _read_workbook_rows(xlsx_path: Path) -> dict[str, dict[str, Any]]:
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Missing dependency: openpyxl. Install it with: pip install openpyxl"
        ) from exc

    workbook = openpyxl.load_workbook(filename=str(xlsx_path), data_only=True)
    sheets: dict[str, dict[str, Any]] = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0

        rows: list[list[Any]] = []
        last_nonempty_row = 0
        last_nonempty_col = 0

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True),
            start=1,
        ):
            row_values = list(row)
            rows.append(row_values)

            row_last_nonempty = 0
            for col_index, value in enumerate(row_values, start=1):
                if value is not None:
                    row_last_nonempty = col_index
            if row_last_nonempty:
                last_nonempty_row = row_index
                last_nonempty_col = max(last_nonempty_col, row_last_nonempty)

        trimmed_rows = rows[:last_nonempty_row]
        trimmed_rows = [r[:last_nonempty_col] for r in trimmed_rows]
        trimmed_rows_json = [[_json_compatible(v) for v in r] for r in trimmed_rows]

        sheet_payload: dict[str, Any] = {
            "rows": trimmed_rows_json,
        }

        header_row_index = None
        for i, row in enumerate(trimmed_rows[:50]):
            if _looks_like_header_row(row):
                header_row_index = i
                break

        if header_row_index is not None:
            header_row = trimmed_rows[header_row_index]
            headers_raw: list[str] = []
            for col_index, cell in enumerate(header_row, start=1):
                if isinstance(cell, str) and cell.strip():
                    headers_raw.append(cell.strip())
                else:
                    headers_raw.append(f"column_{col_index}")
            headers = _make_unique_headers(headers_raw)

            records: list[dict[str, Any]] = []
            for row in trimmed_rows[header_row_index + 1 :]:
                record: dict[str, Any] = {}
                for header, value in zip(headers, row, strict=False):
                    record[header] = _json_compatible(value)
                records.append(record)
            sheet_payload["headers"] = headers
            sheet_payload["records"] = records

        sheets[sheet_name] = sheet_payload

    return sheets


def convert_excel_to_json(input_path: Path, output_path: Path) -> None:
    sheets = _read_workbook_rows(input_path)
    payload = {
        "source_file": str(input_path.resolve()),
        "sheets": sheets,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> int:
    script_dir = Path(__file__).resolve().parent
    default_input = script_dir / "WebsiteShippingRates.xlsx"
    for candidate in ("WebsiteShippingRates.xlsx", "Rates.xlsx"):
        candidate_path = script_dir / candidate
        if candidate_path.exists():
            default_input = candidate_path
            break
    default_output = default_input.with_suffix(".json")

    parser = argparse.ArgumentParser(description="Convert an Excel .xlsx file into JSON.")
    parser.add_argument(
        "input",
        nargs="?",
        default=str(default_input),
        help="Path to the .xlsx file (default: a .xlsx next to this script).",
    )
    parser.add_argument(
        "output",
        nargs="?",
        default=str(default_output),
        help="Path to the .json output file (default: same name as input, .json).",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    convert_excel_to_json(input_path=input_path, output_path=output_path)
    print(f"Wrote JSON: {output_path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
